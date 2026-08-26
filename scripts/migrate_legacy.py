#!/usr/bin/env python3
"""Convierte el histórico Excel legado en una copia restaurable de PalletFlow.

Uso: python scripts/migrate_legacy.py origen.xlsx salida.json [informe.json]
La salida real debe guardarse en .private/ (ignorada por git).
"""
from __future__ import annotations

import hashlib
import json
import re
import sys
import uuid
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

NS = uuid.UUID("7f41ea8d-6619-4ecb-a84c-f7ad8589665c")

def uid(value: str) -> str:
    return str(uuid.uuid5(NS, value))

def clean(value) -> str:
    return str(value or "").strip()

def date_iso(value) -> str:
    text = clean(value)
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    raise ValueError(f"Fecha no reconocida: {text!r}")

def time_iso(value, fallback_date: str) -> str:
    text = clean(value)
    for fmt in ("%d/%m/%Y, %H:%M:%S", "%d/%m/%Y %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).replace(tzinfo=timezone.utc).isoformat().replace("+00:00", "Z")
        except ValueError:
            pass
    return f"{fallback_date}T12:00:00Z"

def rows(ws):
    max_row = ws.max_row
    if ws.tables:
        _, _, _, max_row = range_boundaries(next(iter(ws.tables.values())).ref)
    values = ws.iter_rows(min_row=1, max_row=max_row, values_only=True)
    headers = [clean(v) for v in next(values)]
    for raw in values:
        row = {headers[i]: raw[i] if i < len(raw) else None for i in range(len(headers))}
        if any(v not in (None, "") for v in raw):
            yield row

def base(entity_id: str, timestamp: str):
    return {"id": entity_id, "createdAt": timestamp, "updatedAt": timestamp, "deletedAt": None, "version": 1}

def main(source: Path, target: Path, report_target: Path):
    workbook = load_workbook(source, read_only=False, data_only=True)
    summary = list(rows(workbook["Resumen"]))
    pallet_rows = list(rows(workbook["Palés"]))
    parcel_rows = list(rows(workbook["Packing list"]))

    reception_by_key = defaultdict(list)
    receptions = []
    for index, row in enumerate(summary, 1):
        reference, received = clean(row["Contenedor"]), date_iso(row["Fecha"])
        rid = uid(f"reception:{index}:{reference}:{received}")
        timestamp = f"{received}T08:00:00Z"
        expected = int(row["Palés"] or 0)
        reception = {**base(rid, timestamp), "reference": reference, "vehicle": reference,
            "receivedAt": received, "expectedPallets": max(1, expected), "notes": "Importado del histórico 26/08/2026",
            "status": "descarga"}
        receptions.append(reception)
        reception_by_key[(reference, received)].append((reception, int(row["Grabados"] or 0)))

    # Cuando el resumen contiene dos descargas con igual referencia/fecha, las filas
    # históricas pertenecen a la que tiene grabaciones; la vacía se conserva aparte.
    parcel_groups = defaultdict(list)
    for row in parcel_rows:
        key = (clean(row["Contenedor"]), date_iso(row["Fecha"]), clean(row["Palé"]).upper())
        parcel_groups[key].append(row)

    coord_counts = Counter(clean(r["Ubicación rack escaneada"] or r["Hueco del maestro"]).upper() for r in pallet_rows)
    duplicate_coords = {code for code, count in coord_counts.items() if code and count > 1}
    claimed_coords = set()
    location_entities = {}
    pallets, parcels, movements = [], [], []
    used_pallet_codes = set()
    invalid_units = 0

    for row_index, row in enumerate(pallet_rows, 2):
        reference, received = clean(row["Contenedor"]), date_iso(row["Fecha descarga"])
        candidates = reception_by_key[(reference, received)]
        row_recorded = bool(clean(row["Grabado"])) or int(row["Bultos"] or 0) > 0
        reception = (max(candidates, key=lambda item: item[1]) if row_recorded else min(candidates, key=lambda item: item[1]))[0]
        label = clean(row["Palé"]).upper()
        match = re.search(r"(\d+)", label)
        number = int(match.group(1)) if match else row_index - 1
        code = clean(row["Codificación QR"]) or f"P{number} {reference} {received}"
        if code in used_pallet_codes:
            code = f"{code}-R{receptions.index(reception)+1}"
        used_pallet_codes.add(code)
        pid = uid(f"pallet:{reception['id']}:{number}:{code}")
        timestamp = time_iso(row["Grabado"], received)
        boxes = parcel_groups[(reference, received, label)] if row_recorded else []
        valid_units = sum(int(b["Uds"] or 0) for b in boxes if int(b["Uds"] or 0) > 0)
        invalid_units += sum(1 for b in boxes if int(b["Uds"] or 0) <= 0)
        mocacotas = [clean(b["Mocacota"]) for b in boxes if clean(b["Mocacota"])]
        article = mocacotas[0] if mocacotas else clean(row["Artículo (mocacota)"])
        coord = clean(row["Ubicación rack escaneada"] or row["Hueco del maestro"]).upper()
        assign_coord = bool(coord) and coord not in claimed_coords
        if assign_coord:
            claimed_coords.add(coord)
            lid = uid(f"location:{coord}")
            location_entities.setdefault(coord, {**base(lid, timestamp), "code": coord, "zone": "LEGACY", "aisle": clean(row["Pasillo"]),
                "module": clean(row["Número"]), "level": clean(row["Altura"]), "slot": clean(row["Número"]), "capacity": 1,
                "multiple": False, "status": "ocupada"})
        else:
            lid = None
        recorded = bool(boxes) or bool(clean(row["Grabado"]))
        notes = clean(row["Observaciones"])
        if coord in duplicate_coords and not assign_coord:
            notes = (notes + f" | CONFLICTO LEGADO: ubicación duplicada {coord}; pendiente de revisar").strip(" |")
        pallet = {**base(pid, timestamp), "code": code, "number": number, "receptionId": reception["id"],
            "status": "ubicado" if lid else ("escaneado" if recorded else "pendiente"), "article": article,
            "mocacota": article or None, "expectedParcels": len(boxes) or None, "scanClosedAt": timestamp if recorded else None,
            "parcels": len(boxes), "units": valid_units, "tone": clean(row["Tono"]) or None, "notes": notes,
            "operator": clean(row["Operario"]) or "Importación", "locationId": lid, "locatedAt": timestamp if lid else None}
        pallets.append(pallet)
        movements.append({**base(uid(f"movement:create:{pid}"), timestamp), "palletId": pid, "type": "creacion",
            "reason": "Importado del histórico legado", "operator": pallet["operator"], "deviceId": "legacy-import-20260826"})
        for box_index, box in enumerate(boxes, 1):
            units = int(box["Uds"] or 0)
            btime = time_iso(box["Hora"], received)
            bid = uid(f"parcel:{pid}:{box_index}:{clean(box['Código bulto'])}")
            parcels.append({**base(bid, btime), "code": clean(box["Código bulto"]), "palletId": pid,
                "article": clean(box["Mocacota"]), "mocacota": clean(box["Mocacota"]) or None, "units": max(0, units),
                "operator": clean(box["Operario"]) or pallet["operator"], "anomalous": bool(clean(box["Anomalía"])) or units <= 0})

    backup = {"format": "palletflow-lab-backup", "version": 1, "createdAt": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
        "receptions": receptions, "pallets": pallets, "parcels": parcels, "locations": list(location_entities.values()), "movements": movements}
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text(json.dumps(backup, ensure_ascii=False, indent=2), encoding="utf-8")
    digest = hashlib.sha256(source.read_bytes()).hexdigest()
    report = {"source": str(source), "sha256": digest, "receptions": len(receptions), "pallets": len(pallets),
        "parcels": len(parcels), "units": sum(p["units"] for p in parcels), "locations": len(location_entities),
        "duplicateLocationGroups": len(duplicate_coords), "palletsPendingLocationReview": sum(coord_counts[c]-1 for c in duplicate_coords),
        "invalidUnitRows": invalid_units, "target": str(target)}
    report_target.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))

if __name__ == "__main__":
    if len(sys.argv) < 3:
        raise SystemExit("Uso: migrate_legacy.py origen.xlsx salida.json [informe.json]")
    source, target = Path(sys.argv[1]), Path(sys.argv[2])
    report = Path(sys.argv[3]) if len(sys.argv) > 3 else target.with_suffix(".report.json")
    main(source, target, report)
